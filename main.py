"""Program for creating flexible schedules"""

from abc import ABC, abstractmethod
from random import randint, choice, shuffle, random
import datetime
import json
import tkinter as tk
from tkinter import ttk
from os.path import isfile
from openpyxl import load_workbook


class Main:
    """Main window of program"""

    def __init__(self):
        self.main_window = tk.Tk()
        self.main_window.resizable(0, 0)
        self.pleasures_frame = PleasuresFrame(self, "Удовольствия")
        self.schedule_frame = ScheduleFrame(self, "Расписание")
        self.routines_frame = RoutinesFrame(self, "Дела")
        self.go_button = tk.Button(self.main_window, text='Вперёд!', command=self.__make_schedule,
                                   height=2, width=91)
        self.go_button.pack(side=tk.BOTTOM, anchor=tk.S)
        self.textbox = tk.Text(self.main_window)
        self.textbox.pack()
        self.main_window.mainloop()

    def __make_schedule(self):
        """Make a schedule based on options"""
        self.textbox.delete('1.0', tk.END)
        routines = list(get_routines().values())
        routines.sort(key=lambda x: len(x['active_work_blocks']))
        work_blocks = get_work_blocks()
        for work_block in work_blocks:
            work_block['routines'] = []
        for routine in routines:
            active_work_blocks = routine['active_work_blocks']
            if active_work_blocks:
                index = choice(active_work_blocks)
                work_block = dict(work_blocks[index])
                work_block['routines'].append(routine)
                work_blocks[index] = work_block

        schedule_list = get_schedule() + work_blocks
        schedule_list.sort(key=lambda x: x['start'])
        for schedule_paragraph in schedule_list:
            end_minute = schedule_paragraph['end']
            start_minute = schedule_paragraph['start']
            if 'duration' in schedule_paragraph.keys():  # work block
                self.insert_work_block(schedule_paragraph, start_minute, end_minute)
            else:  # just a paragraph
                self.insert_paragraph(schedule_paragraph, start_minute, end_minute)
        for pleasure in get_pleasures().values():
            if random() * 100 > pleasure['probability']:
                self.textbox.insert(tk.END, "Удовольствие [{}] запрещено\n".format(pleasure['name']))

    def insert_paragraph(self, schedule_paragraph, start_minute, end_minute):
        """Insert paragraph data to textbox"""
        name = schedule_paragraph['name']
        start_str = minutes_to_time(start_minute)
        if start_minute != end_minute:
            end_str = minutes_to_time(end_minute)
            self.textbox.insert(tk.END, "{} - {} {}\n".format(start_str, end_str, name))
        else:
            self.textbox.insert(tk.END, "{} {}\n".format(start_str, name))

    def insert_work_block(self, schedule_paragraph, start_minute, end_minute):
        """Insert work block data to textbox"""
        minutes_of_work = schedule_paragraph['duration']
        sequence = schedule_paragraph['routines']
        minutes_of_rest = end_minute - start_minute - minutes_of_work
        for routine in sequence:
            minutes_of_work -= routine['duration']
        while minutes_of_work >= 45:
            time_block = choice(range(45, min(120, minutes_of_work) + 1, 15))
            sequence.append({'name': "Цикл работы", 'duration': time_block})
            minutes_of_work -= time_block
        minutes_of_rest += minutes_of_work
        shuffle(sequence)
        sequence.append({'name': "Отдых", 'duration': minutes_of_rest})

        for paragraph in sequence:
            start_str = minutes_to_time(start_minute)
            name = paragraph['name']
            if paragraph['duration'] == 0:
                self.textbox.insert(tk.END, "{} {}\n".format(start_str, name))
            else:
                # we created start_str so we can change start_minute
                start_minute += paragraph['duration']
                end_str = minutes_to_time(start_minute)
                self.textbox.insert(tk.END, "{} - {} {}\n".format(start_str, end_str, name))


class Frame(ABC):

    def __init__(self, main: Main, name: str):
        self.name = name
        self.main = main
        self.frame = tk.LabelFrame(main.main_window, text=name)
        self.listbox = tk.Listbox(self.frame, width=35, font=("Courier", 10))
        self.redact_button = None
        self.delete_button = None
        self.pack()

    @abstractmethod
    def pack(self):
        self.fill_listbox()
        self.frame.pack(side=tk.LEFT, anchor=tk.N, fill=tk.Y)
        self.listbox.pack(side=tk.TOP)

    @abstractmethod
    def fill_listbox(self):
        pass

    def update(self):
        self.listbox.delete(0, tk.END)
        self.fill_listbox()

    def activate_disabled_buttons(self):
        self.redact_button.configure(state=tk.NORMAL)
        self.delete_button.configure(state=tk.NORMAL)

    def disable_buttons(self):
        self.redact_button.configure(state=tk.DISABLED)
        self.delete_button.configure(state=tk.DISABLED)


class PleasuresFrame(Frame):

    def pack(self):
        super().pack()
        self.listbox.bind("<Button-1>", lambda _: self.activate_disabled_buttons())
        add_pleasure = tk.Button(self.frame, text="Добавить удовольствие",
                                 command=lambda: PleasureGetter(self.main).pack())
        self.redact_button = tk.Button(self.frame, text="Редактировать", state=tk.DISABLED,
                                       command=self.change_pleasure_window)
        self.delete_button = tk.Button(self.frame, text="Удалить", state=tk.DISABLED,
                                       command=self.delete_pleasure)
        add_pleasure.pack(side=tk.BOTTOM)
        self.redact_button.pack(side=tk.BOTTOM)
        self.delete_button.pack(side=tk.BOTTOM)

    def fill_listbox(self):
        pleasures_dictionary = get_pleasures()
        for pleasure in pleasures_dictionary:
            pleasure_object = Pleasure(pleasures_dictionary[pleasure])
            self.listbox.insert(tk.END, pleasure_object.get_string())

    def change_pleasure_window(self):
        dictionary = create_pleasure_dict_by_string(self.listbox.get(tk.ACTIVE))
        PleasureGetter(self.main, **dictionary).pack()

    def delete_pleasure(self):
        data = get_json_data()
        dictionary = create_pleasure_dict_by_string(self.listbox.get(tk.ACTIVE))
        self.listbox.delete(tk.ACTIVE)
        self.disable_buttons()
        data['pleasures'].pop(dictionary['name'])
        write_json_data(data)


class ScheduleFrame(Frame):

    def pack(self):
        """Create schedule frame"""
        super().pack()
        self.listbox.bind("<Button-1>", lambda _: self.activate_disabled_buttons())
        self.redact_button = tk.Button(self.frame, text="Редактировать", state=tk.DISABLED,
                                       command=self.redact_schedule)
        self.delete_button = tk.Button(self.frame, text="Удалить", state=tk.DISABLED,
                                       command=self.delete_paragraph_or_work_block)
        add_schedule_paragraph = tk.Button(self.frame, text="Добавить пункт плана",
                                           command=lambda: ParagraphGetter(self.main).pack())
        add_work_block = tk.Button(self.frame, text="Добавить блок работы",
                                   command=lambda: WorkBlockGetter(self.main).pack())
        self.frame.pack(side=tk.LEFT, anchor=tk.N, fill=tk.Y)
        add_schedule_paragraph.pack(side=tk.BOTTOM)
        add_work_block.pack(side=tk.BOTTOM)
        self.redact_button.pack(side=tk.BOTTOM)
        self.delete_button.pack(side=tk.BOTTOM)

    def fill_listbox(self):
        schedule_list = get_schedule() + get_work_blocks()
        schedule_list.sort(key=lambda x: x['start'])
        for paragraph in schedule_list:
            if 'duration' in paragraph.keys():
                work_block = WorkBlock(paragraph)
                self.listbox.insert(tk.END, work_block.get_string())
            else:
                paragraph = Paragraph(paragraph)
                self.listbox.insert(tk.END, paragraph.get_string())

    def redact_schedule(self):
        dictionary = create_work_block_or_paragraph_dict_by_string(self.listbox.get(tk.ACTIVE))
        if 'duration' in dictionary:
            WorkBlockGetter(self.main, **dictionary).pack()
        else:
            ParagraphGetter(self.main, **dictionary).pack()

    def delete_paragraph_or_work_block(self):
        data = get_json_data()
        dictionary = create_work_block_or_paragraph_dict_by_string(self.listbox.get(tk.ACTIVE))
        self.listbox.delete(tk.ACTIVE)
        self.disable_buttons()
        if 'duration' in dictionary:
            data['work_blocks'].remove(dictionary)
        else:
            data['schedule'].remove(dictionary)
        write_json_data(data)


class RoutinesFrame(Frame):
    def pack(self):
        self.listbox['width'] = 38
        self.listbox.bind("<Button-1>", lambda _: self.activate_disabled_buttons())
        self.redact_button = tk.Button(self.frame, text="Редактировать", state=tk.DISABLED,
                                       command=self.redact_routine)
        self.delete_button = tk.Button(self.frame, text="Удалить", state=tk.DISABLED,
                                       command=self.delete_routine)
        self.fill_listbox()
        add_routine = tk.Button(self.frame, text="Добавить дело",
                                command=lambda: RoutineGetter(self.main).pack())
        self.frame.pack(side=tk.LEFT, anchor=tk.N, fill=tk.Y)
        self.listbox.pack()
        add_routine.pack(side=tk.BOTTOM)
        self.redact_button.pack(side=tk.BOTTOM)
        self.delete_button.pack(side=tk.BOTTOM)

    def fill_listbox(self):
        routines_dict = get_routines()
        for routine in routines_dict.keys():
            self.listbox.insert(tk.END, Routine(routines_dict[routine]).get_string())

    def redact_routine(self):
        dictionary = create_routine_dict_by_string(self.listbox.get(tk.ACTIVE))
        RoutineGetter(self.main, **dictionary).pack()

    def delete_routine(self):
        data = get_json_data()
        dictionary = create_routine_dict_by_string(self.listbox.get(tk.ACTIVE))
        self.listbox.delete(tk.ACTIVE)
        self.disable_buttons()
        data['routines'].pop(dictionary['name'])
        write_json_data(data)


class ObjectFrame(ABC):

    @abstractmethod
    def get_string(self):
        pass


class Routine(ObjectFrame):
    """Some thing that sometimes should be done, like sports or cleaning"""

    def __init__(self, dictionary):
        self.name = dictionary["name"]
        self.duration = dictionary["duration"]
        self.active_work_blocks = dictionary["active_work_blocks"]

    def get_string(self) -> str:
        pluses_minuses = ''
        for index in range(len(get_work_blocks())):
            if index in self.active_work_blocks:
                pluses_minuses += '+'
            else:
                pluses_minuses += '-'
        return "{:23s}[{}] {}".format(self.name, minutes_to_time(self.duration), pluses_minuses)


class Pleasure(ObjectFrame):
    """Pleasure which can be forbidden for the sake of dopamine quality"""

    def __init__(self, dictionary: dict):
        self.name = dictionary['name']
        self.probability = tk.IntVar(value=dictionary['probability'])
        self.probability.trace('w', lambda *args: self.__update_json())

    def get_probability(self) -> int:
        """Get probability in entry"""
        try:
            probability = self.probability.get()
        except tk.TclError:
            probability = 0
        return probability

    def dictionary(self):
        """Get data about pleasure as dictionary"""
        return {"name": self.name, "probability": self.get_probability()}

    def __update_json(self):
        """Update pleasure in json"""
        pleasures = get_pleasures()
        pleasures[self.name] = self.dictionary()
        write_pleasures(pleasures)

    def get_string(self):
        return "{:30s} {:3d}%".format(self.name, self.probability.get())


class Paragraph(ObjectFrame):
    """Paragraph of schedule"""

    def __init__(self, dictionary):
        self.name = dictionary["name"]
        self.start = dictionary["start"]
        self.end = dictionary["end"]
        self.text = self.get_string()

    def dictionary(self):
        return {"name": self.name, "start": self.start, "end": self.end}

    def get_string(self) -> str:
        """Get string with paragraph info"""
        start_str = minutes_to_time(self.start)
        if self.start != self.end:
            end_str = minutes_to_time(self.end)
            return "{} - {} {}".format(start_str, end_str, self.name)
        return "{} {}".format(start_str, self.name)


class WorkBlock(ObjectFrame):
    """
    Work block
    schedule paragraph which will be randomly filled with routines and work times"""

    def __init__(self, dictionary):
        self.start = dictionary['start']
        self.end = dictionary['end']
        self.duration = dictionary['duration']
        self.text = self.get_string()

    def get_string(self) -> str:
        """Get a string with work block info"""
        duration = minutes_to_time(self.duration)
        start_str = minutes_to_time(self.start)
        if self.start != self.end:
            end_str = minutes_to_time(self.end)
            return "{} - {} Блок работы [{}]".format(start_str, end_str, duration)
        return "{} Блок работы [{}]".format(start_str, duration)


class NameGetter:
    """Entry for name of something"""

    def __init__(self, window, name=''):
        self.name_frame = tk.Frame(window)
        self.name_label = tk.Label(self.name_frame, text="Название")
        self.name_entry = tk.Entry(self.name_frame)
        if name:
            self.name_entry.insert(tk.END, name)

    def pack(self):
        """Pack the frame"""
        self.name_frame.pack()
        self.name_label.pack(side=tk.LEFT)
        self.name_entry.pack(side=tk.LEFT)

    def name(self):
        """Get the name"""
        return self.name_entry.get()


class TimeGetter:
    """Entry for time of something"""

    def __init__(self, window, text, time=0):
        minutes = time % 60
        hours = time // 60
        minutes_list = list(range(0, 60, 5))
        hours_list = list(range(0, 25))
        self.frame = tk.Frame(window)
        self.label = tk.Label(self.frame, text=text)
        self.hours_entry = ttk.Combobox(self.frame, values=hours_list, width=2)
        self.hours_entry.current(hours_list.index(hours))  # choose zero as a default hour
        self.minutes_entry = ttk.Combobox(self.frame, values=minutes_list, width=2)
        self.minutes_entry.current(minutes_list.index(minutes))  # choose zero as a default minute

    def pack(self):
        """Pack frame"""
        self.frame.pack()
        self.label.pack(side=tk.LEFT)
        self.hours_entry.pack(side=tk.LEFT)
        tk.Label(self.frame, text=':').pack(side=tk.LEFT)
        self.minutes_entry.pack(side=tk.LEFT)

    def time(self):
        """Get time"""
        return int(self.hours_entry.get()) * 60 + int(self.minutes_entry.get())


class NumberGetter:
    """Entry for getting a number"""
    def __init__(self, window, text, default_number=''):
        self.frame = tk.Frame(window)
        self.label = tk.Label(self.frame, text=text)
        self.duration_entry = tk.Entry(self.frame, width=3)
        if default_number:
            self.duration_entry.insert(tk.END, default_number)

    def pack(self):
        """Pack a frame"""
        self.frame.pack()
        self.label.pack(side=tk.LEFT)
        self.duration_entry.pack(side=tk.LEFT)

    def time(self):
        """Get time"""
        return int(self.duration_entry.get())


class ObjectGetter(ABC):
    """Window with entries for objects like pleasures and work blocks"""

    def __init__(self, master: Main):
        self.master = master
        self.window = tk.Tk()
        self.okay_button = tk.Button(self.window, text="OK",
                                     command=self.append_to_json)

    @abstractmethod
    def pack(self):
        """Create a window with entries"""
        self.okay_button.pack()
        self.window.mainloop()

    @abstractmethod
    def append_to_json(self):
        """Append data about object to json-file"""

    @abstractmethod
    def paragraph(self):
        """Get data about object as dictionary"""


class PleasureGetter(ObjectGetter):
    """Window with entries for pleasure"""

    def __init__(self, master: Main, name='', probability=''):
        super().__init__(master)
        self.name_frame = NameGetter(self.window, name)
        self.probability = NumberGetter(self.window, "Вероятность", probability)

    def pack(self):
        self.name_frame.pack()
        self.probability.pack()
        super().pack()

    def append_to_json(self):
        """Write data about pleasure to json"""
        paragraph = self.paragraph()
        data = get_json_data()
        data['pleasures'][self.name()] = paragraph
        write_json_data(data)
        self.master.pleasures_frame.update()
        self.window.destroy()

    def paragraph(self) -> dict:
        """Get data about pleasure as dictionary"""
        return {'name': self.name_frame.name(), 'probability': self.probability.time()}

    def name(self):
        """Get pleasure's name"""
        return self.name_frame.name()


class ParagraphGetter(ObjectGetter):
    """Window with entries for schedule paragraph"""

    def __init__(self, master: Main, name='', start=0, end=0):
        super().__init__(master)
        self.old_paragraph = {"name": name, "start": start, "end": end}
        self.name_frame = NameGetter(self.window, name)
        self.start_frame = TimeGetter(self.window, "Начало", start)
        self.end_frame = TimeGetter(self.window, "Конец", end)

    def pack(self):
        """Create a window and run it"""
        self.name_frame.pack()
        self.start_frame.pack()
        self.end_frame.pack()
        super().pack()

    def append_to_json(self):
        """Write data about schedule paragraph into json"""
        paragraph = self.paragraph()
        data = get_json_data()
        if self.old_paragraph in data['schedule']:
            data['schedule'].remove(self.old_paragraph)
        data['schedule'].append(paragraph)
        write_json_data(data)
        self.master.schedule_frame.update()
        self.window.destroy()

    def paragraph(self) -> dict:
        """Get paragraph properties as dictionary"""
        return {"name": self.name_frame.name(),
                "start": self.start_frame.time(),
                "end": self.end_frame.time()}


class RoutineGetter(ObjectGetter):
    """Window with entries for routine"""

    def __init__(self, master: Main, name='', duration=0, active_work_blocks=None):
        super().__init__(master)
        if active_work_blocks is None:
            active_work_blocks = []
        self.name_frame = NameGetter(self.window, name)
        self.duration_frame = TimeGetter(self.window, "Длительность", duration)
        self.active_work_blocks = active_work_blocks

        self.bottom_frame = tk.Frame(self.window)
        self.off_frame = tk.LabelFrame(self.bottom_frame, text="Дело не попадёт в эти блоки работы")
        self.off_listbox = tk.Listbox(self.off_frame, width=40)

        self.on_frame = tk.LabelFrame(self.bottom_frame, text="Дело попадёт в одно из этих блоков работы")
        self.on_listbox = tk.Listbox(self.on_frame, width=40)

    def pack(self):
        """Create a new window and run it"""
        self.name_frame.pack()
        self.duration_frame.pack()
        self.bottom_frame.pack(side=tk.TOP)
        self.off_frame.pack(side=tk.LEFT)
        self.off_listbox.pack()
        self.off_listbox.bind("<Button-1>", lambda _: self.from_off_listbox_to_on())
        self.on_frame.pack(side=tk.RIGHT)
        self.on_listbox.pack()
        self.on_listbox.bind("<Button-1>", lambda _: self.from_on_listbox_to_off())
        self.fill_listboxes()
        super().pack()

    def append_to_json(self):
        """Write data about routine to json"""
        paragraph = self.paragraph()
        data = get_json_data()
        data['routines'][self.name()] = paragraph
        write_json_data(data)
        self.master.routines_frame.update()
        self.window.destroy()

    def paragraph(self) -> dict:
        """Get routine properties as dictionary"""
        return {"name": self.name(),
                "duration": self.duration_frame.time(),
                "active_work_blocks": self.active_work_blocks}

    def fill_listboxes(self):
        work_blocks = get_work_blocks()
        index = 0
        for work_block in work_blocks:
            work_block_object = WorkBlock(work_block)
            work_block_string = '{}. {}'.format(index+1, work_block_object.get_string())
            if index in self.active_work_blocks:
                self.on_listbox.insert(tk.END, work_block_string)
            else:
                self.off_listbox.insert(tk.END, work_block_string)
            index += 1

    def name(self) -> str:
        """Get the name of routine"""
        return self.name_frame.name()

    def from_off_listbox_to_on(self):
        work_block = self.off_listbox.get(tk.ACTIVE)
        if work_block:
            work_block_index = int(work_block.split(sep='.')[0]) - 1  # '2. 14:30 - 21:00 Work block [03:00]' -> 1
            self.active_work_blocks.append(work_block_index)
            self.off_listbox.delete(tk.ACTIVE)
            self.on_listbox.insert(tk.END, work_block)

    def from_on_listbox_to_off(self):
        work_block = self.on_listbox.get(tk.ACTIVE)
        if work_block:
            work_block_index = int(work_block.split(sep='.')[0]) - 1  # '2. 14:30 - 21:00 Work block [03:00]' -> 1
            self.active_work_blocks.remove(work_block_index)
            self.on_listbox.delete(tk.ACTIVE)
            self.off_listbox.insert(tk.END, work_block)


class WorkBlockGetter(ObjectGetter):
    """Window with entries for work block"""

    def __init__(self, master: Main, start=0, end=0, duration=0):
        super().__init__(master)
        self.old_work_block = {"start": start, "end": end, "duration": duration}
        self.start_frame = TimeGetter(self.window, "Начало", start)
        self.end_frame = TimeGetter(self.window, "Конец", end)
        self.duration_frame = TimeGetter(self.window, "Длительность", duration)

    def pack(self):
        """Create a new window and run it"""
        self.start_frame.pack()
        self.end_frame.pack()
        self.duration_frame.pack()
        super().pack()

    def append_to_json(self):
        """Write data about work block into json"""
        paragraph = self.paragraph()
        data = get_json_data()
        if self.old_work_block in data['work_blocks']:
            data['work_blocks'].remove(self.old_work_block)
        data['work_blocks'].append(paragraph)
        data['work_blocks'].sort(key=lambda x: x['start'])
        write_json_data(data)
        self.master.schedule_frame.update()
        self.master.routines_frame.update()
        self.window.destroy()

    def paragraph(self) -> dict:
        """Get work block properties as dictionary"""
        return {"start": self.start_frame.time(),
                "end": self.end_frame.time(),
                "duration": self.duration_frame.time()}


def minutes_to_time(minutes: int) -> str:
    """Converts minutes to format hh:mm"""
    return f'{minutes//60:0>2}:{minutes%60:0>2}'


def time_to_minutes(time: str) -> int:
    """Converts format hh:mm to minutes"""
    list_of_numbers = time.split(sep=':')
    return int(list_of_numbers[0])*60 + int(list_of_numbers[1])


def routines_choose(number):
    """This function is now used now but will be used in future. Don't touch"""
    routines = dict()  # key: name of the routine. value: weigh
    for _ in range(number):
        weigh_sum = sum(routines.values())
        for routine in routines:
            if randint(1, weigh_sum) <= routines[routine]:
                print(routine)
                routines.pop(routine)
                break
            weigh_sum -= routines[routine]


def display_schedule_table(schedule, forbidden_pleasures, day_tuple):
    """This function is now used now but will be used in future. Don't touch"""
    day_today = datetime.datetime(*day_tuple)
    file_name = '../Наработки/Планы.xlsx'
    workbook = load_workbook(file_name)
    worksheet = workbook['Дневной']
    column = get_table_column(worksheet, day_today)
    row = 2
    time = 390
    for routine in schedule:
        cell = worksheet.cell(column=column, row=row)
        routine.table_paragraph(cell, time)
        row += 1
        time += routine.duration

    for pleasure in forbidden_pleasures:
        cell = worksheet.cell(column=column, row=row)
        cell.value = "Нельзя " + pleasure
        row += 1

    workbook.save(file_name)


def get_table_column(worksheet, day_today):
    """This function is now used now but will be used in future. Don't touch"""
    column = 1
    while True:
        cell = worksheet.cell(row=1, column=column)
        if cell.value is None:
            cell.value = day_today
            break
        if cell.value == day_today:
            break
        column += 1
    return column


def create_pleasure_dict_by_string(string: str) -> dict:
    dictionary = dict()
    list_of_words = string.split()
    dictionary["name"] = ' '.join(list_of_words[:-1])  # ['Junk', 'Food', '5%'] -> 'Junk Food'
    dictionary["probability"] = int(list_of_words[-1][:-1])  # ['Junk', 'Food', '5%'] -> 5
    return dictionary


def create_work_block_or_paragraph_dict_by_string(string: str) -> dict:
    dictionary = dict()
    list_of_words = string.split()
    dictionary['start'] = time_to_minutes(list_of_words[0])
    list_of_words = list_of_words[1:]  # Removing first word which is a start time
    if list_of_words[0] == '-':
        dictionary['end'] = time_to_minutes(list_of_words[1])
        list_of_words = list_of_words[2:]  # Removing '-' and end time
    else:
        dictionary['end'] = dictionary['start']
    if "Блок работы" in string:
        dictionary['duration'] = time_to_minutes(list_of_words[-1][1:-1])  # Work block [hh:mm] -> minutes
    else:
        dictionary['name'] = ' '.join(list_of_words)
    return dictionary


def create_routine_dict_by_string(string: str) -> dict:
    dictionary = dict()
    list_of_words = string.split()
    dictionary['name'] = ' '.join(list_of_words[:-2])
    dictionary['duration'] = time_to_minutes(list_of_words[-2][1:-1])  # 'A B C [01:30] +++' -> 90
    dictionary['active_work_blocks'] = []
    for index in range(len(list_of_words[-1])):
        if list_of_words[-1][index] == '+':
            dictionary['active_work_blocks'].append(index)
    return dictionary


def get_json_data():
    """Get all the data from json"""
    with open('data.json', 'r', encoding='utf-8') as file:
        return json.load(file)


def get_pleasures() -> dict:
    """Get pleasures"""
    return get_json_data()['pleasures']


def get_schedule() -> list:
    """Get schedule without work blocks"""
    return get_json_data()['schedule']


def get_routines() -> dict:
    """Get routines"""
    return get_json_data()['routines']


def get_work_blocks() -> list:
    """Get work blocks"""
    return get_json_data()['work_blocks']


def write_json_data(data) -> None:
    """Update the json"""
    with open('data.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


def write_pleasures(pleasures: dict) -> None:
    """Update pleasures"""
    data = get_json_data()
    data['pleasures'] = pleasures
    write_json_data(data)


if not isfile("data.json"):
    write_json_data({"pleasures": {}, "schedule": [], "work_blocks": [], "routines": {}})

Main()

# TODO Пофиксить баг, который может случиться, если на одноразовую рутину не хватает времени
